#!/usr/bin/env python3
"""Refresh the public substance registry from official Japanese sources.

The generated JSON intentionally keeps the source wording for substance names and
concentration values.  It is a browsing aid, not an automated legal-applicability
decision tool.

Run from the repository root:
    python3 scripts/data/update_substance_registry.py
"""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
import io
import json
from pathlib import Path
import re
import unicodedata
from urllib.request import Request, urlopen
import xml.etree.ElementTree as ET

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
except ImportError as exc:  # pragma: no cover - a clear message for maintainers
    raise SystemExit(
        "openpyxl is required. Install it with `python3 -m pip install openpyxl`."
    ) from exc


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_PATH = REPOSITORY_ROOT / "_data" / "substance_registry.json"

ORDER_API_URL = "https://elaws.e-gov.go.jp/api/1/lawdata/347CO0000000318"
ORGANIC_REGULATION_API_URL = "https://elaws.e-gov.go.jp/api/1/lawdata/347M50002000036"
CONCENTRATION_XLSX_URL = "https://www.mhlw.go.jp/content/11300000/001252610.xlsx"

ORDER_VIEW_URL = "https://laws.e-gov.go.jp/law/347CO0000000318"
ORGANIC_REGULATION_VIEW_URL = "https://laws.e-gov.go.jp/law/347M50002000036"
SPECIFIED_CHEMICAL_REGULATION_VIEW_URL = "https://laws.e-gov.go.jp/law/347M50002000039"
CONCENTRATION_OVERVIEW_URL = "https://www.mhlw.go.jp/stf/newpage_32871.html"
CONFIRMATION_MEASUREMENT_URL = (
    "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000046255_00002.html"
)

JST = timezone(timedelta(hours=9))

# 有機則第1条の区分。更新時に条文の記載が変われば、下の検証で停止して見直せる。
FIRST_ORGANIC_SOLVENT_ITEM_NUMBERS = {28, 38}
SECOND_ORGANIC_SOLVENT_ITEM_NUMBERS = (
    set(range(1, 14))
    | set(range(15, 23))
    | {24, 25, 30, 34, 35, 37}
    | set(range(39, 43))
    | set(range(44, 48))
)

COMMON_SEARCH_REPLACEMENTS = (
    ("アルフア", "アルファ"),
    ("フア", "ファ"),
    ("フイ", "フィ"),
    ("フエ", "フェ"),
    ("フオ", "フォ"),
    ("ニツケル", "ニッケル"),
    ("弗化", "フッ化"),
    ("弗素", "フッ素"),
    ("沃化", "ヨウ化"),
    ("沃素", "ヨウ素"),
    ("砒", "ヒ"),
)


def fetch_bytes(url: str) -> bytes:
    request = Request(
        url,
        headers={"User-Agent": "rodoeisei-lab-substance-registry-updater/1.0"},
    )
    with urlopen(request, timeout=60) as response:  # noqa: S310 - fixed public URLs
        return response.read()


def text_content(element: ET.Element | None) -> str:
    if element is None:
        return ""

    # e-Gov represents furigana as <Ruby>砒<Rt>ひ</Rt></Ruby>.  The public
    # registry should show the legal spelling (砒素), while searches can still
    # use the visible substance name naturally.
    parts: list[str] = []

    def visit(node: ET.Element) -> None:
        if node.tag != "Rt":
            if node.text:
                parts.append(node.text)
            for child in node:
                visit(child)
        if node.tail:
            parts.append(node.tail)

    visit(element)
    return "".join(parts).strip()


def find_table(root: ET.Element, title: str) -> ET.Element:
    for table in root.iter("AppdxTable"):
        if text_content(table.find("AppdxTableTitle")) == title:
            return table
    raise RuntimeError(f"e-Gov data did not contain {title}.")


def japanese_number_label(element: ET.Element, child_tag: str) -> str:
    return text_content(element.find(child_tag))


def substance_name_aliases(name: str) -> list[str]:
    """Add common contemporary spellings without changing official display text."""
    aliases = {name, name.replace("―", "ー"), name.replace("―", "-")}
    for source, replacement in COMMON_SEARCH_REPLACEMENTS:
        aliases.update(alias.replace(source, replacement) for alias in tuple(aliases))
    return sorted(alias for alias in aliases if alias)


def organic_solvent_records(order_root: ET.Element, organic_rule_root: ET.Element) -> list[dict[str, object]]:
    organic_rule_text = text_content(organic_rule_root)
    expected_rule_phrases = (
        "第二十八号又は第三十八号",
        "第一号から第十三号まで",
        "第四十四号から第四十七号まで",
    )
    missing_phrases = [
        phrase for phrase in expected_rule_phrases if phrase not in organic_rule_text
    ]
    if missing_phrases:
        raise RuntimeError(
            "The organic-solvent category mapping may have changed. "
            f"Review 有機則第1条 before updating: {missing_phrases}"
        )

    table = find_table(order_root, "別表第六の二")
    records: list[dict[str, object]] = []
    for item in table.findall("Item"):
        item_number = int(item.attrib["Num"])
        name = text_content(item.find("ItemSentence"))
        if not name or name == "削除":
            continue

        if item_number in FIRST_ORGANIC_SOLVENT_ITEM_NUMBERS:
            category = "第1種有機溶剤"
        elif item_number in SECOND_ORGANIC_SOLVENT_ITEM_NUMBERS:
            category = "第2種有機溶剤"
        else:
            category = "第3種有機溶剤"

        official_number = japanese_number_label(item, "ItemTitle")
        records.append(
            {
                "id": f"organic-solvent-{item_number:02d}",
                "system": "有機溶剤中毒予防規則（有機則）",
                "system_key": "organic-solvent",
                "category": category,
                "name": name,
                "law_basis": f"労働安全衛生法施行令 別表第六の二 第{official_number}号",
                "source_url": ORDER_VIEW_URL,
                "source_label": "e-Gov：安衛令 別表第六の二",
                "classification_url": ORGANIC_REGULATION_VIEW_URL,
                "classification_label": "e-Gov：有機則 第1条（区分）",
                "status": "current",
                "status_label": "現行",
                "guidance": "有機則上の区分です。法令の適用は、含有率、作業内容、作業場所なども確認します。",
                "search_text": " ".join(
                    [*substance_name_aliases(name), category, "有機則", "有機溶剤"]
                ),
            }
        )
    if len(records) < 40:
        raise RuntimeError(f"Only {len(records)} organic-solvent entries were parsed.")
    return records


def specified_chemical_records(order_root: ET.Element) -> list[dict[str, object]]:
    table = find_table(order_root, "別表第三")
    records: list[dict[str, object]] = []

    for group in table.findall("Item"):
        group_title = text_content(group.find("ItemSentence"))
        if group_title not in {"第一類物質", "第二類物質", "第三類物質"}:
            continue
        group_number = group.attrib["Num"]

        for item in group.findall("Subitem1"):
            name = text_content(item.find("Subitem1Sentence"))
            # Each class ends with an inclusive mixture clause.  It is important
            # context, but is not a named substance record for this first release.
            if not name or ("掲げる物を" in name and "製剤その他の物" in name):
                continue
            official_number = japanese_number_label(item, "Subitem1Title")
            records.append(
                {
                    "id": f"specified-chemical-{group_number}-{item.attrib['Num']}",
                    "system": "特定化学物質障害予防規則（特化則）",
                    "system_key": "specified-chemical",
                    "category": group_title,
                    "name": name,
                    "law_basis": f"労働安全衛生法施行令 別表第三 第{group_number}号 {official_number}",
                    "source_url": ORDER_VIEW_URL,
                    "source_label": "e-Gov：安衛令 別表第三",
                    "classification_url": SPECIFIED_CHEMICAL_REGULATION_VIEW_URL,
                    "classification_label": "e-Gov：特化則",
                    "status": "current",
                    "status_label": "現行",
                    "guidance": "特化則上の類別です。作業・製剤・含有率・適用除外などを含め、最新の法令で確認します。",
                    "search_text": " ".join(
                        [
                            *substance_name_aliases(name),
                            group_title,
                            "特化則",
                            "特定化学物質",
                        ]
                    ),
                }
            )
    if len(records) < 50:
        raise RuntimeError(f"Only {len(records)} specified-chemical entries were parsed.")
    return records


def merged_cell_values(worksheet: object) -> dict[tuple[int, int], object]:
    values: dict[tuple[int, int], object] = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, column)] = anchor
    return values


def worksheet_value(worksheet: object, merged_values: dict[tuple[int, int], object], row: int, column: int) -> object:
    value = worksheet.cell(row, column).value
    return value if value is not None else merged_values.get((row, column))


def display_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_application_date(value: object) -> tuple[date | None, str]:
    raw = display_text(value)
    if not raw:
        return None, ""
    if isinstance(value, datetime):
        return value.date(), raw
    if isinstance(value, date):
        return value, raw
    if isinstance(value, (int, float)):
        parsed = from_excel(value)
        if isinstance(parsed, datetime):
            parsed_date = parsed.date()
            return parsed_date, f"{parsed_date.year}年{parsed_date.month}月{parsed_date.day}日"
        if isinstance(parsed, date):
            return parsed, f"{parsed.year}年{parsed.month}月{parsed.day}日"

    normalized = unicodedata.normalize("NFKC", raw)
    match = re.search(r"令和\s*(\d+)年\s*(\d+)月\s*(\d+)日", normalized)
    if match:
        year, month, day = (int(value) for value in match.groups())
        return date(2018 + year, month, day), raw
    return None, raw


def application_date_details(value: object, today: date) -> dict[str, str]:
    effective_date, source_text = parse_application_date(value)
    if effective_date is None:
        return {
            "application_date": "",
            "application_date_source": source_text,
            "status": "unknown",
            "status_label": "適用期日を公式一覧で確認",
        }

    rendered = f"{effective_date.year}年{effective_date.month}月{effective_date.day}日施行"
    if effective_date > today:
        rendered += "予定"
        status = "upcoming"
    else:
        status = "current"
    return {
        "application_date": rendered,
        "application_date_source": source_text,
        "status": status,
        "status_label": rendered if status == "upcoming" else "施行済み",
    }


def concentration_standard_records(xlsx_data: bytes, today: date) -> tuple[list[dict[str, object]], str]:
    workbook = load_workbook(io.BytesIO(xlsx_data), data_only=True)
    try:
        worksheet = workbook["濃度基準値等"]
    except KeyError as exc:
        raise RuntimeError("The official workbook no longer contains the 濃度基準値等 sheet.") from exc

    source_snapshot = display_text(worksheet.cell(2, 8).value)
    merged_values = merged_cell_values(worksheet)
    grouped: dict[tuple[str, str, str, str, str, str], dict[str, object]] = {}

    for row in range(4, worksheet.max_row + 1):
        name = display_text(worksheet_value(worksheet, merged_values, row, 1))
        if name == "備考":
            break
        eight_hour = display_text(worksheet_value(worksheet, merged_values, row, 3))
        short_term = display_text(worksheet_value(worksheet, merged_values, row, 4))
        if not name or (not eight_hour and not short_term):
            continue

        cas_rn = display_text(worksheet_value(worksheet, merged_values, row, 2))
        sampling_method = display_text(worksheet_value(worksheet, merged_values, row, 5))
        analysis_method = display_text(worksheet_value(worksheet, merged_values, row, 6))
        application = application_date_details(
            worksheet_value(worksheet, merged_values, row, 8), today
        )
        key = (
            name,
            eight_hour,
            short_term,
            sampling_method,
            analysis_method,
            application["application_date"],
        )
        record = grouped.setdefault(
            key,
            {
                "name": name,
                "cas_rns": [],
                "eight_hour_value": eight_hour,
                "short_term_value": short_term,
                "sampling_method": sampling_method,
                "analysis_method": analysis_method,
                **application,
            },
        )
        if cas_rn and cas_rn not in record["cas_rns"]:
            record["cas_rns"].append(cas_rn)

    records: list[dict[str, object]] = []
    for index, record in enumerate(grouped.values(), start=1):
        records.append(
            {
                "id": f"concentration-standard-{index:03d}",
                "system": "濃度基準値",
                "system_key": "concentration-standard",
                "category": "濃度基準値設定物質",
                "law_basis": "労働安全衛生規則第577条の2第2項に基づく濃度基準値等一覧",
                "source_url": CONCENTRATION_XLSX_URL,
                "source_label": "厚生労働省：濃度基準値等一覧（Excel）",
                "classification_url": CONFIRMATION_MEASUREMENT_URL,
                "classification_label": "厚生労働省：確認測定の考え方",
                "guidance": "濃度基準値を確認する物質です。確認測定を含むばく露状況の確認方法は、作業条件に応じて検討します。",
                "search_text": " ".join(
                    part
                    for part in [
                        *substance_name_aliases(record["name"]),
                        " ".join(record["cas_rns"]),
                        record["eight_hour_value"],
                        record["short_term_value"],
                        "濃度基準値 確認測定 リスクアセスメント",
                    ]
                    if part
                ),
                **record,
            }
        )
    if len(records) < 200:
        raise RuntimeError(f"Only {len(records)} concentration-standard entries were parsed.")
    return records, source_snapshot


def main() -> int:
    today = datetime.now(JST).date()
    order_root = ET.fromstring(fetch_bytes(ORDER_API_URL))
    organic_rule_root = ET.fromstring(fetch_bytes(ORGANIC_REGULATION_API_URL))
    concentration_records, concentration_snapshot = concentration_standard_records(
        fetch_bytes(CONCENTRATION_XLSX_URL), today
    )

    records = [
        *organic_solvent_records(order_root, organic_rule_root),
        *specified_chemical_records(order_root),
        *concentration_records,
    ]
    payload = {
        "schema_version": 1,
        "generated_at": today.isoformat(),
        "scope": [
            "有機溶剤中毒予防規則（有機則）の区分",
            "特定化学物質障害予防規則（特化則）に関係する安衛令別表第三の物質",
            "濃度基準値設定物質",
        ],
        "notes": [
            "この一覧は制度別の確認入口です。個別の法令適用は、物質名、含有率、作業内容、作業場所、設備、適用除外等を最新の一次情報で確認してください。",
            "濃度基準値等一覧のCAS RNは参考情報です。対象物質の当否は、CAS RNではなく物質名に該当するかで確認してください。",
            "濃度基準値設定物質を、個人ばく露測定の一律の対象物質リストとして扱っていません。",
        ],
        "sources": [
            {
                "label": "e-Gov：労働安全衛生法施行令",
                "url": ORDER_VIEW_URL,
                "used_for": "別表第三、別表第六の二",
            },
            {
                "label": "e-Gov：有機溶剤中毒予防規則",
                "url": ORGANIC_REGULATION_VIEW_URL,
                "used_for": "第1条の有機溶剤区分",
            },
            {
                "label": "e-Gov：特定化学物質障害予防規則",
                "url": SPECIFIED_CHEMICAL_REGULATION_VIEW_URL,
                "used_for": "特化則の確認",
            },
            {
                "label": "厚生労働省：濃度基準値等一覧",
                "url": CONCENTRATION_XLSX_URL,
                "used_for": concentration_snapshot or "濃度基準値設定物質",
            },
            {
                "label": "厚生労働省：濃度基準値",
                "url": CONCENTRATION_OVERVIEW_URL,
                "used_for": "告示・関連資料",
            },
        ],
        "records": records,
    }
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(
        f"Wrote {OUTPUT_PATH.relative_to(REPOSITORY_ROOT)} with {len(records)} records "
        f"({len(concentration_records)} concentration-standard records)."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
